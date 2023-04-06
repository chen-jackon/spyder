import requests
from bs4 import BeautifulSoup
import pandas as pd
import folium
import re
import time
import random
import openpyxl
import xlrd

class wuhanSpyer(object):
    def __init__(self):
        self.url = "http://wjw.wuhan.gov.cn/ztzl_28/fk/yqtb/index_{}.shtml"
        # 添加userAgent列表
        self.user_agent_list = [
            "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; WOW64) Gecko/20100101 Firefox/61.0",
            "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.62 Safari/537.36",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
            "Mozilla/5.0 (Macintosh; U; PPC Mac OS X 10.5; en-US; rv:1.9.2.15) Gecko/20110303 Firefox/3.6.15",
        ]

        self.connect = "http://wjw.wuhan.gov.cn/ztzl_28/fk/yqtb"
        self.realUrl = []

    def getSourceData(self):
        # 解析武汉市文生健康委员会网站获取武汉本地的疫情通报
        for i in range(200):  # 最多200页
            print(i)
            if i == 0:
                url = "http://wjw.wuhan.gov.cn/ztzl_28/fk/yqtb/index.shtml"  # 武汉卫健委官网
            else:
                url = "http://wjw.wuhan.gov.cn/ztzl_28/fk/yqtb/index_{}.shtml".format(i)  # 第二页开始遍历
            BASE_URL = url
            response = requests.get(BASE_URL,
                                    headers={'User-Agent': random.choice(self.user_agent_list)})  # 随机切换useragent 增大容错率
            if response.status_code == 200:
                context = response.content.decode("utf-8")
                tmp = re.findall('script[\s\S]*url.*=.*"\.(.*?)"[\s\S]*var[\s\S]*武汉',
                                 context, flags=0)
                self.realUrl.append(self.connect + str(*tmp))  #
        with open("./realUrlList.txt", 'w', encoding="utf-8") as f:  # 将连接写入文本文件中
            for i, n in enumerate(self.realUrl):
                f.write(n)
                f.write('\n')

    def match(self):
        with open("text", "r", encoding="utf-8") as f:
            context = f.read()
        # print(context)
        context="abcabcabc"
        tmp = re.findall('script[\s\S]*url.*=.*"\.(.*?)"[\s\S]*var[\s\S]*武汉', context)
        # tmp = re.findall(, context)

        print(tmp)

    def parseExcel(self):
        res_dict = {}
        path = "./yiqing.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet1 = wb["Sheet1"]
        max_colom=sheet1.max_column
        max_row=sheet1.max_row
        for i in range(max_row-2):
            cell=sheet1.cell(i+2, 2).value
            date = sheet1.cell(i+2, 1).value
            timeArray = time.localtime((date-25569) * 86400.0)
            timeStr = time.strftime('%Y-%m-%d', timeArray)
            quezhen = re.findall('.*确诊.*居住于(.*)', cell)
            wuzhengzhuang = re.findall('.*无症状.*居住于(.*).*', cell)
            res_dict[timeStr] = {"quezhen":quezhen,
                              "wuzhengzhuang":wuzhengzhuang}
            # res["date"] = {"quezhen":quezhen,
            #                "wuzhengzhuang":wuzhengzhuang}
        print(res_dict)

a = wuhanSpyer()
a.parseExcel()
